from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import requests
import json
import math
import concurrent.futures
import pandas as pd
from datetime import datetime
import time

all_FMP_OrdersList = "All_FMP_Orders.csv"
fmpMasterFile = "C:/Users/scifi/OneDrive/Desktop/FMP_MASTER_DATA_FILE_29_10_2022.xlsx"
squareFootFile = "C:/Users/scifi/OneDrive/Desktop/Size Square foot.xlsx"

def generateMonthlySalesReport():
    columns_to_use = ['Order Date', 'SKU', 'Quantity']
    df = pd.read_csv(all_FMP_OrdersList, usecols=columns_to_use)
    df['Order Date'] = pd.to_datetime(df['Order Date'])
    df['Month_Year'] = df['Order Date'].dt.strftime('%B %Y')
    pivot_table = df.pivot_table(
        index='SKU', columns='Month_Year', values='Quantity', aggfunc='sum', fill_value=0)
    latest_month_year = datetime.now().strftime('%B %Y')
    month_year_order = [latest_month_year] + \
        [month_year for month_year in df['Month_Year'].unique() if month_year !=
         latest_month_year]
    pivot_table = pivot_table[month_year_order]
    pivot_table['Total'] = pivot_table.sum(axis=1)
    pivot_table = pivot_table.sort_values(by='Total', ascending=False)
    pivot_table.reset_index(inplace=True)

    # Create an Excel workbook and add the DataFrame to a sheet
    wb = Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
            cell.alignment = Alignment(wrap_text=True)

    # Autofit column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the Excel workbook as XLSX
    wb.save('C:/Users/scifi/OneDrive/Desktop/FMP Reports/Sales_Report_Each_Month_Sorted_By_Total.xlsx')

    print(f'Sales Report for Each Month Sorted by Total Generated: Sales_Report_Each_Month_Sorted_By_Total.xlsx')


def generateMonthlySalesReportAlternating():
    columns_to_use = ['Order Date', 'SKU', 'Quantity']
    df = pd.read_csv(all_FMP_OrdersList, usecols=columns_to_use)
    df['Order Date'] = pd.to_datetime(df['Order Date'])
    df['Month_Year'] = df['Order Date'].dt.strftime('%B %Y')

    # Get unique months and sort them in descending order
    months = df['Month_Year'].unique()
    months = sorted(months, key=lambda x: datetime.strptime(
        x, '%B %Y'), reverse=True)

    # Create a dictionary to store quantities for each month
    month_quantities = {}
    for month in months:
        month_quantities[month] = {}

    for month in months:
        # Get the corresponding quantities for the month
        month_df = df[df['Month_Year'] == month].groupby(
            'SKU')['Quantity'].sum().reset_index()

        # Sort the SKUs by quantity for the current month
        month_df = month_df.sort_values(by='Quantity', ascending=False)

        # Store the SKUs and quantities in the dictionary
        month_quantities[month]['SKUs'] = month_df['SKU'].tolist()
        month_quantities[month]['Quantities'] = month_df['Quantity'].tolist()

    # Create a Pandas DataFrame with alternating SKU and Quantity columns
    final_columns = []

    for month in months:
        final_columns.extend([f"{month} SKU", f"{month} Quantity", ''])

    final_df = pd.DataFrame(columns=final_columns)

    # Calculate the maximum number of SKUs among all months
    max_length = max(len(month_quantities[month]['SKUs']) for month in months)

    for i in range(max_length):
        row_data = []
        for month in months:
            if i < len(month_quantities[month]['SKUs']):
                row_data.append(month_quantities[month]['SKUs'][i])
                row_data.append(month_quantities[month]['Quantities'][i])
                row_data.append('')
            else:
                row_data.extend(['', '', ''])
        final_df.loc[len(final_df)] = row_data

    # Create an Excel workbook and add the DataFrame to a sheet
    wb = Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
            cell.alignment = Alignment(wrap_text=True)

    # Autofit column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the Excel workbook
    wb.save('C:/Users/scifi/OneDrive/Desktop/FMP Reports/Monthly_Sales_Report_Alternating_Months.xlsx')
    print(f'Sales Report For Each Month Individually SKUs Sorted Generated: Monthly_Sales_Report_Alternating_Months.xlsx')


def generateCombinedSalesReport(order_data_file, product_data_file, square_foot_file, output_file):
    # Load the product master data
    product_data = pd.read_excel(product_data_file)

    # Load the order data
    order_data = pd.read_csv(order_data_file)

    # Convert "Order Date" column to datetime
    order_data['Order Date'] = pd.to_datetime(order_data['Order Date'])

    # Extract month and year from "Order Date"
    order_data['Year'] = order_data['Order Date'].dt.year
    order_data['Month'] = order_data['Order Date'].dt.strftime('%B %Y')

    order_data["SKU"] = order_data["SKU"].str.upper()
    product_data["ProductID"] = product_data["ProductID"].str.upper()

    # Merge the order data with the product master data using SKU as the key
    merged_data = order_data.merge(
        product_data, left_on="SKU", right_on="ProductID", how="left")

    # Group and sum the Quantity by SKU, Group, Color, Size, Month, and Year
    result_data = merged_data.groupby(["SKU", "ProductGroupName", "COLOR", "SIZE", "Month"])[
        "Quantity"].sum().reset_index()

    # Pivot the data to have separate columns for each month
    result_data = result_data.pivot_table(index=["SKU", "ProductGroupName", "COLOR", "SIZE"],
                                          columns="Month", values="Quantity", fill_value=0).reset_index()

    # Sort the columns chronologically
    sorted_columns = sorted(result_data.columns[4:], key=lambda x: pd.to_datetime(
        x, format='%B %Y'), reverse=True)
    result_data = result_data[[
        'SKU', 'ProductGroupName', 'COLOR', 'SIZE'] + sorted_columns]

    # Rename columns for the final result
    result_data.columns = ["SKU", "Group", "Color",
                           "Size"] + result_data.columns[4:].tolist()

    # Standardize the "Size" column format
    result_data['Size'] = result_data['Size'].str.upper()
    result_data['Size'] = result_data['Size'].str.replace("X", " X ")
    result_data['Size'] = result_data['Size'].str.replace("  X  ", " X ")
    result_data['Size'] = result_data['Size'].str.replace(
        "HE X AGON", "HEXAGON")

    # Load the Square Foot data
    square_foot_data = pd.read_excel(
        square_foot_file, usecols=["Size", "SQ-FT"])

    square_foot_data["Size"] = square_foot_data["Size"].str.upper()

    # Merge the DataFrames on the "Size" column using a left join
    result_data = result_data.merge(
        square_foot_data, left_on="Size", right_on="Size", how="left")

    # Fill missing "SQ-FT" values with 0
    result_data["SQ-FT"].fillna(0, inplace=True)

    # Multiply quantity columns by SQ-FT columns for each month
    for col in sorted_columns:
        result_data[col] = round(result_data[col] * result_data["SQ-FT"], 2)

    # Find the latest Order Date
    latest_order_date = order_data['Order Date'].max()
    latest_order_date_formatted = latest_order_date.strftime(
        '%d-%m-%Y %H:%M %Z')

    # Sum the values of month columns by group and color
    result_data_group_color = result_data.groupby(
        ["Group", "Color"])[sorted_columns].sum().reset_index()

    # Rename the latest month column with the specified format
    latest_month = sorted_columns[0]
    result_data_group_color.rename(columns={
                                   latest_month: f"{latest_month} (As of {latest_order_date_formatted})"}, inplace=True)

    # Create an Excel writer object and write the result data to a new sheet in an XLSX file
    with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
        result_data_group_color.to_excel(
            writer, sheet_name="Square Foot Report", index=False)

        # Get the xlsxwriter workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets["Square Foot Report"]

        # Adjust the column widths to fit the content
        for i, column in enumerate(result_data_group_color.columns):
            column_len = max(result_data_group_color[column].astype(
                str).str.len().max(), len(column))
            worksheet.set_column(i, i, column_len + 2)  # +2 for padding

    print(f"Square Foot Report Generated: '{output_file}'")


generateCombinedSalesReport(all_FMP_OrdersList, fmpMasterFile, squareFootFile,
                            "C:/Users/scifi/OneDrive/Desktop/FMP Reports/Group_Colour_Square_Foot_Report.xlsx")
generateMonthlySalesReport()
generateMonthlySalesReportAlternating()
exit()

url = "https://au.api.sellercloud.com/rest/api/token"

payload = json.dumps({
    "Username": "shivani.fmp@furnishmyplace.com",
    "Password": "SHIV@ni0101^"
})
headers = {'Content-Type': 'application/json'}

response = requests.request("POST", url, headers=headers, data=payload).json()

token = response["access_token"]
print(token)

allProducts = []
pageNumber = 1
lastPageNumber = 1
total = 1


def getLastPageNumber(pageNumber):
    url = f"https://au.api.sellercloud.com/rest/api/orders?pageNumber={pageNumber}&pageSize=1"

    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer ' + token
    }

    response = requests.request("GET", url, headers=headers,
                                data=payload).json()

    if "Items" in response:
        lastPageNumber = math.ceil(response["TotalResults"] / 50)
        return lastPageNumber, response["TotalResults"]
    else:
        return 0


def getProducts(pageNumber):
    start = time.time()
    global allProducts
    url = f"https://au.api.sellercloud.com/rest/api/orders?pageNumber={pageNumber}&pageSize=50"

    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Bearer ' + token
    }

    response = requests.request("GET", url, headers=headers,
                                data=payload).json()

    if "Items" in response:
        fetchedProducts = response["Items"]
        end = time.time()
        timeTaken = end-start
        print(timeTaken)
        remainingTime = (lastPageNumber-pageNumber)*timeTaken/12
        remainingTime = f'{remainingTime//60} minutes {remainingTime%60} seconds'
        print(
            f'Page: {pageNumber} - {len(fetchedProducts)} Orders Fetched. Total Fetched: {len(allProducts)} - ETA: {remainingTime}'
        )
        allProducts += fetchedProducts


lastPageNumber, totalProducts = getLastPageNumber(pageNumber)

print(f'Total Pages: {lastPageNumber}, Total Products: {totalProducts}')

with concurrent.futures.ThreadPoolExecutor() as executor:
    executor.map(getProducts, range(1, lastPageNumber + 1))

print(f'{len(allProducts)} Products Fetched Successfully.')

# #create a empty data
df = pd.DataFrame(columns=[
    "Order Date", "Product ID", "Product Name", "Quantity", "Cost",
    "Sales Amount", "Profit", "Order ID", "Channel", "Order Source ID",
    "Status"
])

# channels = {
#     "68": "Houzz",
#     "50": "Walmart Marketplace",
#     "4": "Amazon",
#     "1": "Ebay",
#     "20": "FBA",
#     "66": "Drop Ship",
#     "27": "Wayfair",
#     "6": "Website"
# }

index = 0
for product in allProducts:
    orderDate = product["TimeOfOrder"]
    orderSourceID = product["OrderSourceOrderID"]
    status = ""
    for item in product["Items"]:
        productID = item["ProductID"].upper()
        channel = product["OrderSource"]
        productName = item["DisplayName"]
        quantity = item["Qty1"]
        cost = item["SiteCost"]
        salesAmount = item["AdjustedSitePrice"]
        orderID = item["OrderID"]
        profit = round(float(salesAmount) - float(cost), 2)
        df.loc[index] = [
            orderDate, productID, productName, quantity, cost, salesAmount,
            profit, orderID, channel, orderSourceID, status
        ]
        index += 1

df.to_csv(all_FMP_OrdersList, index=False)